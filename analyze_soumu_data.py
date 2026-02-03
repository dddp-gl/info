import openpyxl
import csv

def analyze_soumu_excel():
    """総務省Excelファイルのデータ構造を分析"""
    try:
        wb = openpyxl.load_workbook('soumu_municipal_data.xlsx')
        ws = wb.active
        
        print(f"シート名: {ws.title}")
        print(f"最大行: {ws.max_row}")
        print(f"最大列: {ws.max_column}")
        print("\n最初の20行を表示:")
        
        for row_num in range(1, min(21, ws.max_row + 1)):
            row_data = []
            for col_num in range(1, min(11, ws.max_column + 1)):  # 最初の10列
                cell_value = ws.cell(row=row_num, column=col_num).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            print(f"Row {row_num}: {row_data}")
            
    except Exception as e:
        print(f"エラー: {e}")

def extract_proper_soumu_data():
    """総務省データを正しく抽出"""
    try:
        wb = openpyxl.load_workbook('soumu_municipal_data.xlsx')
        ws = wb.active
        
        # CSVファイルを作成
        with open('fa-jp----_soumu_extracted.csv', 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # ヘッダー行
            writer.writerow([
                'Num', 'x1', 'x2', 'x3', 'x4', 'x5', 'x1Name', 'x2Name', 'x3Name', 
                'x4Name', 'x5Name', 'account', 'subArea', 'Population(K)', 'GNI(M$)'
            ])
            
            num = 1
            data_start_row = None
            current_pref = ""
            current_pref_code = "00"
            city_counter = 0
            
            # データの開始行を見つける
            for row_num in range(1, 50):
                cell_a = ws.cell(row=row_num, column=1).value
                cell_b = ws.cell(row=row_num, column=2).value
                
                if (cell_a and '都道府県' in str(cell_a)) or (cell_b and '市区町村' in str(cell_b)):
                    data_start_row = row_num + 1
                    print(f"データ開始行を発見: {data_start_row}")
                    break
            
            if data_start_row:
                for row_num in range(data_start_row, ws.max_row + 1):
                    # 各列の値を取得
                    pref_cell = ws.cell(row=row_num, column=1).value
                    city_cell = ws.cell(row=row_num, column=2).value
                    pop_cell = ws.cell(row=row_num, column=3).value
                    
                    pref_name = str(pref_cell).strip() if pref_cell else ""
                    city_name = str(city_cell).strip() if city_cell else ""
                    population = pop_cell if isinstance(pop_cell, (int, float)) and pop_cell > 0 else 0
                    
                    # 都道府県が変わった場合
                    if pref_name and pref_name != current_pref and pref_name != "None":
                        current_pref = pref_name
                        current_pref_code = f"{int(current_pref_code) + 1:02d}"
                        city_counter = 0
                        
                        # 都道府県行を追加
                        writer.writerow([
                            num, 'jp', current_pref_code, '', '', '', 'Japan', pref_name, '', 
                            '', '', '', '', int(population/1000) if population > 0 else '', ''
                        ])
                        num += 1
                    
                    # 市区町村がある場合
                    if city_name and city_name != "None" and city_name != current_pref:
                        city_counter += 1
                        city_code = f"{city_counter:02d}"
                        
                        # 市区町村行を追加
                        writer.writerow([
                            num, 'jp', current_pref_code, city_code, '', '', 'Japan', current_pref, city_name,
                            '', '', '', '', int(population/1000) if population > 0 else '', ''
                        ])
                        num += 1
            
            print(f"抽出完了: {num-1} 件のレコード")
            return num - 1
            
    except Exception as e:
        print(f"抽出エラー: {e}")
        return 0

if __name__ == "__main__":
    print("=== 総務省Excelファイルの構造分析 ===")
    analyze_soumu_excel()
    
    print("\n=== データ抽出実行 ===")
    total_records = extract_proper_soumu_data()
    print(f"完了: {total_records} 件のレコードを抽出しました。")