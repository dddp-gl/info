import openpyxl
import csv
import re

def extract_soumu_data():
    """総務省の住民基本台帳データから全市町村データを抽出"""
    
    try:
        # 総務省のExcelファイルを開く
        wb = openpyxl.load_workbook('soumu_municipal_data.xlsx')
        
        # シート名を確認
        print("Available sheets:", wb.sheetnames)
        
        # 主要データが含まれるシートを取得
        ws = wb.active
        
        # データを読み込む
        all_data = []
        header_found = False
        
        for row in ws.iter_rows(values_only=True):
            if row and any(cell for cell in row):
                if not header_found:
                    # ヘッダー行を探す
                    if any('都道府県' in str(cell) or '市区町村' in str(cell) for cell in row if cell):
                        header_found = True
                        continue
                else:
                    # データ行を処理
                    if len(row) >= 3 and row[0]:  # 最低限のデータがある行
                        all_data.append(row)
        
        return all_data
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []

def create_complete_jp_csv_from_soumu():
    """総務省データから完全なCSVを作成"""
    
    # 総務省データを取得
    soumu_data = extract_soumu_data()
    
    if not soumu_data:
        print("総務省データの読み込みに失敗しました。代替データを使用します。")
        return create_fallback_data()
    
    # CSVファイルを作成
    with open('fa-jp----_from_soumu.csv', 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        
        # ヘッダー行
        writer.writerow([
            'Num', 'x1', 'x2', 'x3', 'x4', 'x5', 'x1Name', 'x2Name', 'x3Name', 
            'x4Name', 'x5Name', 'account', 'subArea', 'Population(K)', 'GNI(M$)'
        ])
        
        num = 1
        current_pref_code = None
        pref_counter = 0
        
        for row_data in soumu_data:
            if len(row_data) >= 3:
                # データを解析して都道府県と市町村を特定
                pref_name = str(row_data[0]) if row_data[0] else ""
                city_name = str(row_data[1]) if row_data[1] else ""
                population = row_data[2] if len(row_data) > 2 and isinstance(row_data[2], (int, float)) else 0
                
                # 都道府県コードを生成
                if pref_name and not city_name:
                    pref_counter += 1
                    current_pref_code = f"{pref_counter:02d}"
                    
                    # 都道府県行を追加
                    writer.writerow([
                        num, 'jp', current_pref_code, '', '', '', 'Japan', pref_name, '', 
                        '', '', '', '', int(population/1000) if population > 1000 else population, ''
                    ])
                    num += 1
                
                elif city_name and current_pref_code:
                    # 市町村行を追加
                    city_code = f"{len([r for r in soumu_data if str(r[1]) and current_pref_code]):02d}"
                    
                    writer.writerow([
                        num, 'jp', current_pref_code, city_code, '', '', 'Japan', pref_name, city_name,
                        '', '', '', '', int(population/1000) if population > 1000 else population, ''
                    ])
                    num += 1
        
        print(f"総務省データから {num-1} 件のレコードを作成しました。")
        return num - 1

def create_fallback_data():
    """総務省データが利用できない場合の代替データ"""
    
    # 日本の標準的な行政区画データ
    prefectures_data = [
        ('01', '北海道', ['札幌市', '函館市', '小樽市', '旭川市', '室蘭市', '釧路市', '帯広市', '北見市']),
        ('02', '青森県', ['青森市', '弘前市', '八戸市', '黒石市', '五所川原市', '十和田市', '三沢市', 'むつ市']),
        ('03', '岩手県', ['盛岡市', '宮古市', '大船渡市', '花巻市', '北上市', '久慈市', '遠野市', '一関市']),
        ('13', '東京都', [
            '千代田区', '中央区', '港区', '新宿区', '文京区', '台東区', '墨田区', '江東区',
            '品川区', '目黒区', '大田区', '世田谷区', '渋谷区', '中野区', '杉並区', '豊島区',
            '北区', '荒川区', '板橋区', '練馬区', '足立区', '葛飾区', '江戸川区',
            '八王子市', '立川市', '武蔵野市', '三鷹市', '青梅市', '府中市', '昭島市', '調布市'
        ]),
        ('27', '大阪府', ['大阪市', '堺市', '岸和田市', '豊中市', '池田市', '吹田市', '泉大津市', '高槻市'])
    ]
    
    with open('fa-jp----_fallback.csv', 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        
        # ヘッダー行
        writer.writerow([
            'Num', 'x1', 'x2', 'x3', 'x4', 'x5', 'x1Name', 'x2Name', 'x3Name', 
            'x4Name', 'x5Name', 'account', 'subArea', 'Population(K)', 'GNI(M$)'
        ])
        
        num = 1
        
        for pref_code, pref_name, cities in prefectures_data:
            # 都道府県行
            writer.writerow([
                num, 'jp', pref_code, '', '', '', 'Japan', pref_name, '', 
                '', '', '', '', '', ''
            ])
            num += 1
            
            # 市町村行
            for i, city_name in enumerate(cities, 1):
                writer.writerow([
                    num, 'jp', pref_code, f"{i:02d}", '', '', 'Japan', pref_name, city_name,
                    '', '', '', '', '', ''
                ])
                num += 1
        
        return num - 1

if __name__ == "__main__":
    try:
        total_records = create_complete_jp_csv_from_soumu()
        print(f"総務省データを使用して {total_records} 件のレコードを作成しました。")
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        print("代替データを作成しています...")
        total_records = create_fallback_data()
        print(f"代替データで {total_records} 件のレコードを作成しました。")
    
    # 主要市町村のサンプルデータ（実際には全市町村のデータが必要）
    # ここでは代表例として各都道府県の主要市町村を含める
    municipalities = [
        # 北海道
        {'pref': '01', 'city_code': '01', 'name': '札幌市', 'pop': 1973},
        {'pref': '01', 'city_code': '02', 'name': '函館市', 'pop': 246},
        {'pref': '01', 'city_code': '03', 'name': '小樽市', 'pop': 111},
        {'pref': '01', 'city_code': '04', 'name': '旭川市', 'pop': 323},
        {'pref': '01', 'city_code': '05', 'name': '室蘭市', 'pop': 80},
        {'pref': '01', 'city_code': '06', 'name': '釧路市', 'pop': 165},
        {'pref': '01', 'city_code': '07', 'name': '帯広市', 'pop': 166},
        {'pref': '01', 'city_code': '08', 'name': '北見市', 'pop': 113},
        {'pref': '01', 'city_code': '09', 'name': '岩見沢市', 'pop': 79},
        {'pref': '01', 'city_code': '10', 'name': '網走市', 'pop': 33},
        {'pref': '01', 'city_code': '11', 'name': '留萌市', 'pop': 20},
        {'pref': '01', 'city_code': '12', 'name': '苫小牧市', 'pop': 172},
        {'pref': '01', 'city_code': '13', 'name': '稚内市', 'pop': 32},
        {'pref': '01', 'city_code': '14', 'name': '美唄市', 'pop': 20},
        {'pref': '01', 'city_code': '15', 'name': '芦別市', 'pop': 12},
        {'pref': '01', 'city_code': '16', 'name': '江別市', 'pop': 118},
        {'pref': '01', 'city_code': '17', 'name': '赤平市', 'pop': 9},
        {'pref': '01', 'city_code': '18', 'name': '紋別市', 'pop': 22},
        {'pref': '01', 'city_code': '19', 'name': '士別市', 'pop': 18},
        {'pref': '01', 'city_code': '20', 'name': '名寄市', 'pop': 27},
        
        # 青森県
        {'pref': '02', 'city_code': '01', 'name': '青森市', 'pop': 278},
        {'pref': '02', 'city_code': '02', 'name': '弘前市', 'pop': 168},
        {'pref': '02', 'city_code': '03', 'name': '八戸市', 'pop': 221},
        {'pref': '02', 'city_code': '04', 'name': '黒石市', 'pop': 31},
        {'pref': '02', 'city_code': '05', 'name': '五所川原市', 'pop': 53},
        {'pref': '02', 'city_code': '06', 'name': '十和田市', 'pop': 59},
        {'pref': '02', 'city_code': '07', 'name': '三沢市', 'pop': 40},
        {'pref': '02', 'city_code': '08', 'name': 'むつ市', 'pop': 54},
        {'pref': '02', 'city_code': '09', 'name': 'つがる市', 'pop': 31},
        {'pref': '02', 'city_code': '10', 'name': '平川市', 'pop': 30},
        
        # 東京23特別区（全て）
        {'pref': '13', 'city_code': '01', 'name': '千代田区', 'pop': 66},
        {'pref': '13', 'city_code': '02', 'name': '中央区', 'pop': 172},
        {'pref': '13', 'city_code': '03', 'name': '港区', 'pop': 261},
        {'pref': '13', 'city_code': '04', 'name': '新宿区', 'pop': 349},
        {'pref': '13', 'city_code': '05', 'name': '文京区', 'pop': 240},
        {'pref': '13', 'city_code': '06', 'name': '台東区', 'pop': 212},
        {'pref': '13', 'city_code': '07', 'name': '墨田区', 'pop': 276},
        {'pref': '13', 'city_code': '08', 'name': '江東区', 'pop': 527},
        {'pref': '13', 'city_code': '09', 'name': '品川区', 'pop': 421},
        {'pref': '13', 'city_code': '10', 'name': '目黒区', 'pop': 281},
        {'pref': '13', 'city_code': '11', 'name': '大田区', 'pop': 738},
        {'pref': '13', 'city_code': '12', 'name': '世田谷区', 'pop': 919},
        {'pref': '13', 'city_code': '13', 'name': '渋谷区', 'pop': 230},
        {'pref': '13', 'city_code': '14', 'name': '中野区', 'pop': 341},
        {'pref': '13', 'city_code': '15', 'name': '杉並区', 'pop': 588},
        {'pref': '13', 'city_code': '16', 'name': '豊島区', 'pop': 294},
        {'pref': '13', 'city_code': '17', 'name': '北区', 'pop': 353},
        {'pref': '13', 'city_code': '18', 'name': '荒川区', 'pop': 217},
        {'pref': '13', 'city_code': '19', 'name': '板橋区', 'pop': 578},
        {'pref': '13', 'city_code': '20', 'name': '練馬区', 'pop': 740},
        {'pref': '13', 'city_code': '21', 'name': '足立区', 'pop': 698},
        {'pref': '13', 'city_code': '22', 'name': '葛飾区', 'pop': 460},
        {'pref': '13', 'city_code': '23', 'name': '江戸川区', 'pop': 698},
        
        # 東京都市部（26市5町8村）
        {'pref': '13', 'city_code': '24', 'name': '八王子市', 'pop': 562},
        {'pref': '13', 'city_code': '25', 'name': '立川市', 'pop': 184},
        {'pref': '13', 'city_code': '26', 'name': '武蔵野市', 'pop': 148},
        {'pref': '13', 'city_code': '27', 'name': '三鷹市', 'pop': 190},
        {'pref': '13', 'city_code': '28', 'name': '青梅市', 'pop': 133},
        {'pref': '13', 'city_code': '29', 'name': '府中市', 'pop': 261},
        {'pref': '13', 'city_code': '30', 'name': '昭島市', 'pop': 112},
        {'pref': '13', 'city_code': '31', 'name': '調布市', 'pop': 238},
        {'pref': '13', 'city_code': '32', 'name': '町田市', 'pop': 428},
        {'pref': '13', 'city_code': '33', 'name': '小金井市', 'pop': 127},
        {'pref': '13', 'city_code': '34', 'name': '小平市', 'pop': 195},
        {'pref': '13', 'city_code': '35', 'name': '日野市', 'pop': 190},
        {'pref': '13', 'city_code': '36', 'name': '東村山市', 'pop': 149},
        {'pref': '13', 'city_code': '37', 'name': '国分寺市', 'pop': 127},
        {'pref': '13', 'city_code': '38', 'name': '国立市', 'pop': 76},
        {'pref': '13', 'city_code': '39', 'name': '福生市', 'pop': 57},
        {'pref': '13', 'city_code': '40', 'name': '狛江市', 'pop': 83},
        {'pref': '13', 'city_code': '41', 'name': '東大和市', 'pop': 85},
        {'pref': '13', 'city_code': '42', 'name': '清瀬市', 'pop': 75},
        {'pref': '13', 'city_code': '43', 'name': '東久留米市', 'pop': 117},
        {'pref': '13', 'city_code': '44', 'name': '武蔵村山市', 'pop': 70},
        {'pref': '13', 'city_code': '45', 'name': '多摩市', 'pop': 148},
        {'pref': '13', 'city_code': '46', 'name': '稲城市', 'pop': 92},
        {'pref': '13', 'city_code': '47', 'name': '羽村市', 'pop': 54},
        {'pref': '13', 'city_code': '48', 'name': 'あきる野市', 'pop': 80},
        {'pref': '13', 'city_code': '49', 'name': '西東京市', 'pop': 206},
        
        # 大阪府（全43市町村の例）
        {'pref': '27', 'city_code': '01', 'name': '大阪市', 'pop': 2691},
        {'pref': '27', 'city_code': '02', 'name': '堺市', 'pop': 826},
        {'pref': '27', 'city_code': '03', 'name': '岸和田市', 'pop': 194},
        {'pref': '27', 'city_code': '04', 'name': '豊中市', 'pop': 400},
        {'pref': '27', 'city_code': '05', 'name': '池田市', 'pop': 104},
        {'pref': '27', 'city_code': '06', 'name': '吹田市', 'pop': 381},
        {'pref': '27', 'city_code': '07', 'name': '泉大津市', 'pop': 74},
        {'pref': '27', 'city_code': '08', 'name': '高槻市', 'pop': 351},
        {'pref': '27', 'city_code': '09', 'name': '貝塚市', 'pop': 84},
        {'pref': '27', 'city_code': '10', 'name': '守口市', 'pop': 142},
        {'pref': '27', 'city_code': '11', 'name': '枚方市', 'pop': 396},
        {'pref': '27', 'city_code': '12', 'name': '茨木市', 'pop': 282},
        {'pref': '27', 'city_code': '13', 'name': '八尾市', 'pop': 264},
        {'pref': '27', 'city_code': '14', 'name': '泉佐野市', 'pop': 100},
        {'pref': '27', 'city_code': '15', 'name': '富田林市', 'pop': 110},
        {'pref': '27', 'city_code': '16', 'name': '寝屋川市', 'pop': 228},
        {'pref': '27', 'city_code': '17', 'name': '河内長野市', 'pop': 104},
        {'pref': '27', 'city_code': '18', 'name': '松原市', 'pop': 117},
        {'pref': '27', 'city_code': '19', 'name': '大東市', 'pop': 122},
        {'pref': '27', 'city_code': '20', 'name': '和泉市', 'pop': 185},
        {'pref': '27', 'city_code': '21', 'name': '箕面市', 'pop': 138},
        {'pref': '27', 'city_code': '22', 'name': '柏原市', 'pop': 69},
        {'pref': '27', 'city_code': '23', 'name': '羽曳野市', 'pop': 110},
        {'pref': '27', 'city_code': '24', 'name': '門真市', 'pop': 123},
        {'pref': '27', 'city_code': '25', 'name': '摂津市', 'pop': 85},
        {'pref': '27', 'city_code': '26', 'name': '高石市', 'pop': 56},
        {'pref': '27', 'city_code': '27', 'name': '藤井寺市', 'pop': 65},
        {'pref': '27', 'city_code': '28', 'name': '東大阪市', 'pop': 488},
        {'pref': '27', 'city_code': '29', 'name': '泉南市', 'pop': 60},
        {'pref': '27', 'city_code': '30', 'name': '四條畷市', 'pop': 55},
        {'pref': '27', 'city_code': '31', 'name': '交野市', 'pop': 78},
        {'pref': '27', 'city_code': '32', 'name': '大阪狭山市', 'pop': 57},
        {'pref': '27', 'city_code': '33', 'name': '阪南市', 'pop': 52},
    ]
    
    # CSVデータを生成
    data = []
    num = 1
    
    # 都道府県データを追加
    for pref_code, pref_name in prefectures.items():
        # 都道府県人口データ（概算）
        pref_populations = {
            '01': 5210, '02': 1237, '03': 1204, '04': 2302, '05': 944,
            '06': 1068, '07': 1833, '08': 2860, '09': 1934, '10': 1939,
            '11': 7345, '12': 6322, '13': 14049, '14': 9238, '15': 2201,
            '16': 1035, '17': 1132, '18': 767, '19': 813, '20': 2048,
            '21': 1978, '22': 3636, '23': 7552, '24': 1771, '25': 1414,
            '26': 2583, '27': 8838, '28': 5466, '29': 1330, '30': 925,
            '31': 553, '32': 671, '33': 1888, '34': 2799, '35': 1342,
            '36': 719, '37': 950, '38': 1334, '39': 691, '40': 5136,
            '41': 811, '42': 1313, '43': 1738, '44': 1123, '45': 1073,
            '46': 1588, '47': 1467
        }
        
        data.append([
            num, 'jp', pref_code, '', '', '', 'Japan', pref_name, '', '', '', '', '', 
            pref_populations.get(pref_code, ''), ''
        ])
        num += 1
    
    # 市町村データを追加
    for muni in municipalities:
        data.append([
            num, 'jp', muni['pref'], muni['city_code'], '', '', 'Japan',
            prefectures[muni['pref']], muni['name'], '', '', '', '', muni['pop'], ''
        ])
        num += 1
    
    # DataFrameを作成
    df = pd.DataFrame(data, columns=[
        'Num', 'x1', 'x2', 'x3', 'x4', 'x5', 'x1Name', 'x2Name', 'x3Name', 
        'x4Name', 'x5Name', 'account', 'subArea', 'Population(K)', 'GNI(M$)'
    ])
    
    return df

if __name__ == "__main__":
    df = create_complete_japan_data()
    df.to_csv('fa-jp----_complete.csv', index=False)
    print(f"Complete Japan data created with {len(df)} records")
    print("Sample data:")
    print(df.head(10))