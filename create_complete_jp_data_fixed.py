import openpyxl
import csv
import re

def extract_soumu_data():
    """総務省の住民基本台帳データから全市町村データを抽出"""
    
    try:
        # 総務省のExcelファイルを開く
        wb = openpyxl.load_workbook('soumu_municipal_data.xlsx')
        
        # シート名を確認
        print("Available sheets:", wb.sheetnames)
        
        # 主要データが含まれるシートを取得
        ws = wb.active
        
        # データを読み込む
        all_data = []
        header_found = False
        
        for row in ws.iter_rows(values_only=True):
            if row and any(cell for cell in row):
                if not header_found:
                    # ヘッダー行を探す
                    if any('都道府県' in str(cell) or '市区町村' in str(cell) for cell in row if cell):
                        header_found = True
                        continue
                else:
                    # データ行を処理
                    if len(row) >= 3 and row[0]:  # 最低限のデータがある行
                        all_data.append(row)
        
        return all_data
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []

def create_complete_jp_csv_from_soumu():
    """総務省データから完全なCSVを作成"""
    
    # 総務省データを取得
    soumu_data = extract_soumu_data()
    
    if not soumu_data:
        print("総務省データの読み込みに失敗しました。代替データを使用します。")
        return create_fallback_data()
    
    # CSVファイルを作成
    with open('fa-jp----_from_soumu.csv', 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        
        # ヘッダー行
        writer.writerow([
            'Num', 'x1', 'x2', 'x3', 'x4', 'x5', 'x1Name', 'x2Name', 'x3Name', 
            'x4Name', 'x5Name', 'account', 'subArea', 'Population(K)', 'GNI(M$)'
        ])
        
        num = 1
        current_pref_code = None
        pref_counter = 0
        
        for row_data in soumu_data:
            if len(row_data) >= 3:
                # データを解析して都道府県と市町村を特定
                pref_name = str(row_data[0]) if row_data[0] else ""
                city_name = str(row_data[1]) if row_data[1] else ""
                population = row_data[2] if len(row_data) > 2 and isinstance(row_data[2], (int, float)) else 0
                
                # 都道府県コードを生成
                if pref_name and not city_name:
                    pref_counter += 1
                    current_pref_code = f"{pref_counter:02d}"
                    
                    # 都道府県行を追加
                    writer.writerow([
                        num, 'jp', current_pref_code, '', '', '', 'Japan', pref_name, '', 
                        '', '', '', '', int(population/1000) if population > 1000 else population, ''
                    ])
                    num += 1
                
                elif city_name and current_pref_code:
                    # 市町村行を追加
                    city_code = f"{len([r for r in soumu_data if str(r[1]) and current_pref_code]):02d}"
                    
                    writer.writerow([
                        num, 'jp', current_pref_code, city_code, '', '', 'Japan', pref_name, city_name,
                        '', '', '', '', int(population/1000) if population > 1000 else population, ''
                    ])
                    num += 1
        
        print(f"総務省データから {num-1} 件のレコードを作成しました。")
        return num - 1

def create_fallback_data():
    """総務省データが利用できない場合の代替データ"""
    
    # 日本の標準的な行政区画データ
    prefectures_data = [
        ('01', '北海道', ['札幌市', '函館市', '小樽市', '旭川市', '室蘭市', '釧路市', '帯広市', '北見市']),
        ('02', '青森県', ['青森市', '弘前市', '八戸市', '黒石市', '五所川原市', '十和田市', '三沢市', 'むつ市']),
        ('03', '岩手県', ['盛岡市', '宮古市', '大船渡市', '花巻市', '北上市', '久慈市', '遠野市', '一関市']),
        ('13', '東京都', [
            '千代田区', '中央区', '港区', '新宿区', '文京区', '台東区', '墨田区', '江東区',
            '品川区', '目黒区', '大田区', '世田谷区', '渋谷区', '中野区', '杉並区', '豊島区',
            '北区', '荒川区', '板橋区', '練馬区', '足立区', '葛飾区', '江戸川区',
            '八王子市', '立川市', '武蔵野市', '三鷹市', '青梅市', '府中市', '昭島市', '調布市'
        ]),
        ('27', '大阪府', ['大阪市', '堺市', '岸和田市', '豊中市', '池田市', '吹田市', '泉大津市', '高槻市'])
    ]
    
    with open('fa-jp----_fallback.csv', 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        
        # ヘッダー行
        writer.writerow([
            'Num', 'x1', 'x2', 'x3', 'x4', 'x5', 'x1Name', 'x2Name', 'x3Name', 
            'x4Name', 'x5Name', 'account', 'subArea', 'Population(K)', 'GNI(M$)'
        ])
        
        num = 1
        
        for pref_code, pref_name, cities in prefectures_data:
            # 都道府県行
            writer.writerow([
                num, 'jp', pref_code, '', '', '', 'Japan', pref_name, '', 
                '', '', '', '', '', ''
            ])
            num += 1
            
            # 市町村行
            for i, city_name in enumerate(cities, 1):
                writer.writerow([
                    num, 'jp', pref_code, f"{i:02d}", '', '', 'Japan', pref_name, city_name,
                    '', '', '', '', '', ''
                ])
                num += 1
        
        return num - 1

if __name__ == "__main__":
    try:
        total_records = create_complete_jp_csv_from_soumu()
        print(f"総務省データを使用して {total_records} 件のレコードを作成しました。")
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        print("代替データを作成しています...")
        total_records = create_fallback_data()
        print(f"代替データで {total_records} 件のレコードを作成しました。")