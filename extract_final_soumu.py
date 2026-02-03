import openpyxl
import csv
import re

def extract_complete_soumu_data():
    """総務省の住民基本台帳データから全市町村データを正しく抽出"""
    try:
        wb = openpyxl.load_workbook('soumu_municipal_data.xlsx')
        ws = wb.active
        
        # CSVファイルを作成
        with open('fa-jp----_complete_from_soumu.csv', 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # ヘッダー行
            writer.writerow([
                'Num', 'x1', 'x2', 'x3', 'x4', 'x5', 'x1Name', 'x2Name', 'x3Name', 
                'x4Name', 'x5Name', 'account', 'subArea', 'Population(K)', 'GNI(M$)'
            ])
            
            num = 1
            current_pref = ""
            current_pref_code = "00" 
            pref_codes = {}  # 都道府県名からコードへのマッピング
            city_codes = {}  # 都道府県内の市町村カウンター
            
            # データを7行目から開始（0ベースなので6から）
            for row_num in range(7, ws.max_row + 1):
                # 各列の値を取得
                team_code = ws.cell(row=row_num, column=1).value  # A列: 団体コード
                pref_name = ws.cell(row=row_num, column=2).value  # B列: 都道府県名
                city_name = ws.cell(row=row_num, column=3).value  # C列: 市区町村名  
                population = ws.cell(row=row_num, column=6).value  # F列: 総人口
                
                if not pref_name:
                    continue
                
                pref_name = str(pref_name).strip()
                city_name = str(city_name).strip() if city_name else ""
                
                # 全国合計行はスキップ
                if pref_name == "合計":
                    continue
                    
                # 都道府県の処理
                if city_name == "-":  # 都道府県合計行
                    if pref_name not in pref_codes:
                        current_pref_code = f"{len(pref_codes) + 1:02d}"
                        pref_codes[pref_name] = current_pref_code
                        city_codes[pref_name] = 0
                        
                        # 都道府県行を追加
                        pop_k = int(population / 1000) if population and population > 0 else ""
                        writer.writerow([
                            num, 'jp', current_pref_code, '', '', '', 'Japan', pref_name, '', 
                            '', '', '', '', pop_k, ''
                        ])
                        num += 1
                        current_pref = pref_name
                
                # 市区町村の処理
                elif city_name and pref_name in pref_codes:
                    # 札幌市の区など、市の下位組織は除外（要求に応じて調整可能）
                    if '区' in city_name and ('市' in city_name or pref_name == '東京都'):
                        # 東京23特別区は含める、その他の政令市の区は統合
                        if pref_name == '東京都' and any(ward in city_name for ward in ['千代田区', '中央区', '港区', '新宿区', '文京区', '台東区', '墨田区', '江東区', '品川区', '目黒区', '大田区', '世田谷区', '渋谷区', '中野区', '杉並区', '豊島区', '北区', '荒川区', '板橋区', '練馬区', '足立区', '葛飾区', '江戸川区']):
                            # 東京23特別区は個別に記録
                            pass
                        else:
                            # その他の区は市レベルで統合するためスキップ
                            continue
                    
                    city_codes[pref_name] += 1
                    city_code = f"{city_codes[pref_name]:02d}"
                    pref_code = pref_codes[pref_name]
                    
                    # 市区町村行を追加
                    pop_k = int(population / 1000) if population and population > 0 else ""
                    writer.writerow([
                        num, 'jp', pref_code, city_code, '', '', 'Japan', pref_name, city_name,
                        '', '', '', '', pop_k, ''
                    ])
                    num += 1
            
            print(f"総務省データから {num-1} 件のレコードを抽出しました。")
            return num - 1
            
    except Exception as e:
        print(f"抽出エラー: {e}")
        return 0

if __name__ == "__main__":
    total_records = extract_complete_soumu_data()
    print(f"完了: 総務省の公式データから {total_records} 件のレコードを作成しました。")
    print("ファイル名: fa-jp----_complete_from_soumu.csv")